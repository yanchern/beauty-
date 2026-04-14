from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


EU_COUNTRIES = {
    "AT",
    "BE",
    "BG",
    "HR",
    "CY",
    "CZ",
    "DE",
    "DK",
    "EE",
    "ES",
    "FI",
    "FR",
    "GR",
    "HU",
    "IE",
    "IT",
    "LT",
    "LU",
    "LV",
    "MT",
    "NL",
    "PL",
    "PT",
    "RO",
    "SE",
    "SI",
    "SK",
}

UK_TAX_DIVISOR = Decimal("1.2")
UK_TAX_RATE = Decimal("0.2")
IT_TAX_DIVISOR = Decimal("1.22")
IT_TAX_RATE = Decimal("0.22")


@dataclass(frozen=True)
class RuleContext:
    country_code: str


@dataclass(frozen=True)
class RuleSpec:
    rule_id: str
    logic_group: str
    logic_bucket: str
    description: str
    matcher: Callable[[dict[str, str], RuleContext], bool]


@dataclass(frozen=True)
class MetricSpec:
    label: str
    calculator: Callable[["CountryReportData"], Decimal]


@dataclass(frozen=True)
class CountryConfig:
    code: str
    slug: str
    name_zh: str
    title: str
    description: str
    sales_report_label: str
    logic_doc_label: str
    logic_doc_accept: str
    emblem_path: str
    excluded_transaction_types: tuple[str, ...]
    rules: tuple[RuleSpec, ...]
    summary_metrics: tuple[MetricSpec, ...]
    card_metrics: tuple[MetricSpec, ...]


@dataclass(frozen=True)
class MatchResult:
    rule_id: str
    logic_group: str
    logic_bucket: str
    description: str


@dataclass
class MatchedRow:
    source_row_number: int
    rule_id: str
    logic_group: str
    logic_bucket: str
    rule_description: str
    transaction_event_id: str
    activity_period: str
    marketplace: str
    transaction_type: str
    tax_collection_responsibility: str
    tax_reporting_scheme: str
    seller_sku: str
    asin: str
    item_description: str
    qty: str
    total_activity_value_amt_vat_incl: Decimal
    transaction_currency_code: str
    price_of_items_vat_rate_percent: str
    buyer_vat_number: str
    seller_arrival_country_vat_number: str
    sale_depart_country: str
    sale_arrival_country: str
    taxable_jurisdiction: str
    vat_calculation_imputation_country: str
    invoice_url: str


@dataclass
class CountryReportData:
    country: CountryConfig
    matched_rows: list[MatchedRow]
    group_totals: dict[str, Decimal]
    total_sales: Decimal
    sale_only_total: Decimal


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据国家税金逻辑，从 Amazon 销售报告中生成各国独立的税金汇总。"
    )
    parser.add_argument("csv_path", type=Path, help="季度销售报告 CSV 路径")
    parser.add_argument(
        "--country",
        choices=("fr", "es", "uk", "de", "it"),
        default="fr",
        help="国家代码：fr=法国，es=西班牙，uk=英国，de=德国，it=意大利。",
    )
    parser.add_argument(
        "--logic-pdf-path",
        type=Path,
        default=None,
        help="税金逻辑源文件路径，仅用于在汇总页记录来源。",
    )
    parser.add_argument(
        "--output-path",
        type=Path,
        default=None,
        help="输出 Excel 路径；默认在 CSV 同目录生成 *_税金汇总.xlsx。",
    )
    return parser.parse_args()


def decimal_or_zero(raw: str) -> Decimal:
    value = (raw or "").strip()
    if not value:
        return Decimal("0")
    try:
        return Decimal(value.replace(",", ""))
    except InvalidOperation as exc:
        raise ValueError(f"无法解析金额: {raw!r}") from exc


def normalized(row: dict[str, str], key: str) -> str:
    return (row.get(key) or "").strip()


def upper_value(row: dict[str, str], key: str) -> str:
    return normalized(row, key).upper()


def normalized_transaction_type(row: dict[str, str]) -> str:
    raw = upper_value(row, "TRANSACTION_TYPE")
    return re.sub(r"[_\-\s]+", " ", raw).strip()


def is_blank(raw: str) -> bool:
    return raw.strip() == ""


def is_zero(raw: str) -> bool:
    return decimal_or_zero(raw) == 0


def is_non_zero(raw: str) -> bool:
    value = raw.strip()
    return value != "" and not is_zero(value)


def is_blank_or_zero(raw: str) -> bool:
    value = raw.strip()
    return value == "" or is_zero(value)


def is_explicit_zero(raw: str) -> bool:
    value = raw.strip()
    return value != "" and is_zero(value)


def vat_rate_equals(row: dict[str, str], value: str) -> bool:
    return decimal_or_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT")) == decimal_or_zero(value)


def is_seller_responsible(row: dict[str, str]) -> bool:
    return upper_value(row, "TAX_COLLECTION_RESPONSIBILITY") == "SELLER"


def is_marketplace_responsible(row: dict[str, str]) -> bool:
    return upper_value(row, "TAX_COLLECTION_RESPONSIBILITY") == "MARKETPLACE"


def has_buyer_vat(row: dict[str, str]) -> bool:
    return not is_blank(normalized(row, "BUYER_VAT_NUMBER"))


def has_seller_arrival_vat(row: dict[str, str]) -> bool:
    return not is_blank(normalized(row, "SELLER_ARRIVAL_COUNTRY_VAT_NUMBER"))


def country_equals(row: dict[str, str], key: str, country_code: str) -> bool:
    return upper_value(row, key) == country_code


def country_in_eu(row: dict[str, str], key: str) -> bool:
    return upper_value(row, key) in EU_COUNTRIES


def rule_group_total(group_name: str, label: str | None = None) -> MetricSpec:
    default_label = f"{group_name}(含税)" if group_name.endswith("销售额") else f"{group_name}销售额(含税)"
    return MetricSpec(
        label=label or default_label,
        calculator=lambda report, target=group_name: report.group_totals.get(target, Decimal("0")),
    )


def derived_metric(label: str, calculator: Callable[[CountryReportData], Decimal]) -> MetricSpec:
    return MetricSpec(label=label, calculator=calculator)


def sum_metric(label: str, group_names: tuple[str, ...]) -> MetricSpec:
    return MetricSpec(
        label=label,
        calculator=lambda report, groups=group_names: sum(
            (report.group_totals.get(group, Decimal("0")) for group in groups),
            start=Decimal("0"),
        ),
    )


COUNTRY_CONFIGS: dict[str, CountryConfig] = {}


def base_allowed(row: dict[str, str], ctx: RuleContext) -> bool:
    country = COUNTRY_CONFIGS[ctx.country_code]
    return normalized_transaction_type(row) not in country.excluded_transaction_types


def fr_rule_part_2(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and has_buyer_vat(row)
        and country_equals(row, "SALE_DEPART_COUNTRY", "FR")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
    )


def fr_rule_part_3_strict(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "FR")
    )


def fr_rule_part_3_missing(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_blank_or_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "FR")
    )


def es_rule_part_1(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and has_buyer_vat(row)
        and country_equals(row, "SALE_DEPART_COUNTRY", "ES")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
    )


def es_rule_part_2(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "ES")
    )


def es_rule_part_3(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_blank(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "ES")
    )


def uk_rule_marketplace(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        is_marketplace_responsible(row)
        and country_equals(row, "SALE_DEPART_COUNTRY", "GB")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
    )


def uk_rule_zero_rate(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        is_seller_responsible(row)
        and country_equals(row, "SALE_DEPART_COUNTRY", "GB")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
        and vat_rate_equals(row, "0")
    )


def uk_rule_taxed_part_a(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        is_seller_responsible(row)
        and vat_rate_equals(row, "0.2")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "GB")
    )


def uk_rule_taxed_part_b(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        is_seller_responsible(row)
        and vat_rate_equals(row, "0.2")
        and country_equals(row, "SALE_DEPART_COUNTRY", "GB")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
        and not has_seller_arrival_vat(row)
    )


def de_rule_depart_de_non_zero(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and country_equals(row, "SALE_DEPART_COUNTRY", "DE")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
    )


def de_rule_missing_domestic_zero(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_explicit_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and country_equals(row, "SALE_DEPART_COUNTRY", "DE")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "DE")
    )


def de_rule_missing_destination_de(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "DE")
    )


def it_rule_b1(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and has_buyer_vat(row)
        and country_equals(row, "SALE_DEPART_COUNTRY", "IT")
        and country_in_eu(row, "SALE_ARRIVAL_COUNTRY")
    )


def it_rule_b2(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_non_zero(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "IT")
    )


def it_rule_f(row: dict[str, str], ctx: RuleContext) -> bool:
    return (
        base_allowed(row, ctx)
        and is_seller_responsible(row)
        and is_blank(normalized(row, "PRICE_OF_ITEMS_VAT_RATE_PERCENT"))
        and not has_buyer_vat(row)
        and country_in_eu(row, "SALE_DEPART_COUNTRY")
        and country_equals(row, "SALE_ARRIVAL_COUNTRY", "IT")
    )


COUNTRY_CONFIGS.update(
    {
        "fr": CountryConfig(
            code="fr",
            slug="france",
            name_zh="法国",
            title="法国季度申报税金计算",
            description="按已确认的法国口径提取自行缴税订单，保留第 2 部分、第 3 部分严格口径与遗漏订单口径。",
            sales_report_label="法国销售报告 CSV",
            logic_doc_label="法国税金逻辑文件",
            logic_doc_accept=".pdf,application/pdf",
            emblem_path="/emblem_fr.svg",
            excluded_transaction_types=("COMMINGLING BUY", "COMMINGLING SELLER"),
            rules=(
                RuleSpec(
                    rule_id="FR_SELF_TAX_P2",
                    logic_group="自行缴税第2部分",
                    logic_bucket="规则2-1_FR出发_买家税号有效",
                    description="CQ=SELLER；AE 非空且非 0；CA 有值；BP=FR；BQ 属于欧盟国家。",
                    matcher=fr_rule_part_2,
                ),
                RuleSpec(
                    rule_id="FR_SELF_TAX_P3_STRICT",
                    logic_group="自行缴税第3部分",
                    logic_bucket="规则2-2_目的国FR_买家税号空白_AE非零",
                    description="CQ=SELLER；AE 非空且非 0；CA 空白；BP 属于欧盟国家；BQ=FR。",
                    matcher=fr_rule_part_3_strict,
                ),
                RuleSpec(
                    rule_id="FR_SELF_TAX_P3_MISSING",
                    logic_group="自行缴税第3部分",
                    logic_bucket="规则2-2_目的国FR_买家税号空白_AE空白或零_按遗漏订单纳入",
                    description="CQ=SELLER；AE 为空白或为 0；CA 空白；BP 属于欧盟国家；BQ=FR。",
                    matcher=fr_rule_part_3_missing,
                ),
            ),
            summary_metrics=(
                rule_group_total("自行缴税第2部分"),
                rule_group_total("自行缴税第3部分"),
                derived_metric("自行缴税合计销售额(含税)", lambda report: report.total_sales),
                derived_metric("仅SALE交易销售额总额(含税)", lambda report: report.sale_only_total),
            ),
            card_metrics=(
                rule_group_total("自行缴税第2部分", "自行缴税第2部分"),
                rule_group_total("自行缴税第3部分", "自行缴税第3部分"),
                derived_metric("自行缴税合计", lambda report: report.total_sales),
            ),
        ),
        "es": CountryConfig(
            code="es",
            slug="spain",
            name_zh="西班牙",
            title="西班牙季度申报税金计算",
            description="按西班牙计税原则提取自行缴税订单，独立计算第一部分、第二部分、第三部分。",
            sales_report_label="西班牙销售报告 CSV",
            logic_doc_label="西班牙税金逻辑文件",
            logic_doc_accept=".pdf,application/pdf",
            emblem_path="/emblem_es.svg",
            excluded_transaction_types=("COMMINGLING BUY",),
            rules=(
                RuleSpec(
                    rule_id="ES_SELF_TAX_P1",
                    logic_group="自行缴税第一部分",
                    logic_bucket="规则2-1_ES出发_买家税号有效",
                    description="CQ=SELLER；AE 非空且非 0；CA 有值；BP=ES；BQ 属于欧盟国家。",
                    matcher=es_rule_part_1,
                ),
                RuleSpec(
                    rule_id="ES_SELF_TAX_P2",
                    logic_group="自行缴税第二部分",
                    logic_bucket="规则2-2_目的国ES_买家税号空白_AE非零",
                    description="CQ=SELLER；AE 非空且非 0；CA 空白；BP 属于欧盟国家；BQ=ES。",
                    matcher=es_rule_part_2,
                ),
                RuleSpec(
                    rule_id="ES_SELF_TAX_P3",
                    logic_group="自行缴税第三部分",
                    logic_bucket="规则2-3_目的国ES_买家税号空白_AE空白",
                    description="CQ=SELLER；AE 为空白；CA 空白；BP 属于欧盟国家；BQ=ES。",
                    matcher=es_rule_part_3,
                ),
            ),
            summary_metrics=(
                rule_group_total("自行缴税第一部分"),
                rule_group_total("自行缴税第二部分"),
                rule_group_total("自行缴税第三部分"),
                derived_metric("自行缴税合计销售额(含税)", lambda report: report.total_sales),
                derived_metric("仅SALE交易销售额总额(含税)", lambda report: report.sale_only_total),
            ),
            card_metrics=(
                rule_group_total("自行缴税第一部分", "第一部分"),
                rule_group_total("自行缴税第二部分", "第二部分"),
                rule_group_total("自行缴税第三部分", "第三部分"),
                derived_metric("自行缴税合计", lambda report: report.total_sales),
            ),
        ),
        "uk": CountryConfig(
            code="uk",
            slug="uk",
            name_zh="英国",
            title="英国季度申报税金计算",
            description="按英国税金计算逻辑独立汇总代扣销售额、未代扣 0 税率销售额、未代扣缴税 A/B 部分和应缴税金。",
            sales_report_label="英国销售报告 CSV",
            logic_doc_label="英国税金逻辑 TXT",
            logic_doc_accept=".txt,text/plain",
            emblem_path="/emblem_uk.svg",
            excluded_transaction_types=(),
            rules=(
                RuleSpec(
                    rule_id="UK_MARKETPLACE",
                    logic_group="代扣代缴销售额",
                    logic_bucket="CQ=MARKETPLACE_BP=GB_BQ保留欧盟国家",
                    description="CQ=MARKETPLACE；BP=GB；BQ 保留欧盟国家代码；BA 求和。",
                    matcher=uk_rule_marketplace,
                ),
                RuleSpec(
                    rule_id="UK_ZERO_RATE",
                    logic_group="未代扣0税率销售额",
                    logic_bucket="CQ=SELLER_BP=GB_AE=0",
                    description="CQ=SELLER；BP=GB；BQ 保留欧盟国家代码；AE=0；BA 求和。",
                    matcher=uk_rule_zero_rate,
                ),
                RuleSpec(
                    rule_id="UK_TAXABLE_A",
                    logic_group="未代扣缴税A部分",
                    logic_bucket="CQ=SELLER_AE=0.2_BQ=GB",
                    description="CQ=SELLER；AE=0.2；BQ=GB；BA 求和。",
                    matcher=uk_rule_taxed_part_a,
                ),
                RuleSpec(
                    rule_id="UK_TAXABLE_B",
                    logic_group="未代扣缴税B部分",
                    logic_bucket="CQ=SELLER_AE=0.2_BP=GB_BQ剔除有税号国家及非欧盟国家",
                    description="CQ=SELLER；AE=0.2；BP=GB；BQ 保留欧盟国家代码且卖家目的国税号为空；BA 求和。",
                    matcher=uk_rule_taxed_part_b,
                ),
            ),
            summary_metrics=(
                rule_group_total("代扣代缴销售额"),
                rule_group_total("未代扣0税率销售额"),
                rule_group_total("未代扣缴税A部分"),
                rule_group_total("未代扣缴税B部分"),
                sum_metric("未代扣缴税销售额合计(含税)", ("未代扣缴税A部分", "未代扣缴税B部分")),
                derived_metric(
                    "应缴税金",
                    lambda report: (
                        report.group_totals.get("未代扣缴税A部分", Decimal("0"))
                        + report.group_totals.get("未代扣缴税B部分", Decimal("0"))
                    )
                    / UK_TAX_DIVISOR
                    * UK_TAX_RATE,
                ),
            ),
            card_metrics=(
                rule_group_total("代扣代缴销售额", "代扣销售额"),
                rule_group_total("未代扣0税率销售额", "未代扣0税率"),
                sum_metric("未代扣缴税合计", ("未代扣缴税A部分", "未代扣缴税B部分")),
                derived_metric(
                    "应缴税金",
                    lambda report: (
                        report.group_totals.get("未代扣缴税A部分", Decimal("0"))
                        + report.group_totals.get("未代扣缴税B部分", Decimal("0"))
                    )
                    / UK_TAX_DIVISOR
                    * UK_TAX_RATE,
                ),
            ),
        ),
        "de": CountryConfig(
            code="de",
            slug="germany",
            name_zh="德国",
            title="德国季度申报税金计算",
            description="按德国文档中 Seller 相关口径独立汇总德国发欧盟销售额与两类平台遗漏订单，不额外推算税金。",
            sales_report_label="德国销售报告 CSV",
            logic_doc_label="德国税金逻辑 DOCX",
            logic_doc_accept=".docx,.doc,application/vnd.openxmlformats-officedocument.wordprocessingml.document,application/msword",
            emblem_path="/emblem_de.svg",
            excluded_transaction_types=("COMMINGLING BUY",),
            rules=(
                RuleSpec(
                    rule_id="DE_SELLER_NON_ZERO",
                    logic_group="德国发出欧盟销售额",
                    logic_bucket="CQ=SELLER_AE非零_BP=DE_BQ欧盟",
                    description="CQ=SELLER；AE 非空且非 0；BP=DE；BQ 属于欧盟国家；BA 求和。",
                    matcher=de_rule_depart_de_non_zero,
                ),
                RuleSpec(
                    rule_id="DE_MISSING_DOMESTIC_ZERO",
                    logic_group="平台遗漏订单(德国境内,AE=0)",
                    logic_bucket="CQ=SELLER_AE=0_BP=DE_BQ=DE",
                    description="CQ=SELLER；AE=0；BP=DE；BQ=DE；BA 求和。",
                    matcher=de_rule_missing_domestic_zero,
                ),
                RuleSpec(
                    rule_id="DE_MISSING_DESTINATION",
                    logic_group="平台遗漏订单(目的国德国,AE非零)",
                    logic_bucket="CQ=SELLER_AE非零_CA空白_BP欧盟_BQ=DE",
                    description="CQ=SELLER；AE 非空且非 0；CA 空白；BP 属于欧盟国家；BQ=DE；BA 求和。",
                    matcher=de_rule_missing_destination_de,
                ),
            ),
            summary_metrics=(
                rule_group_total("德国发出欧盟销售额"),
                rule_group_total("平台遗漏订单(德国境内,AE=0)"),
                rule_group_total("平台遗漏订单(目的国德国,AE非零)"),
                derived_metric("Seller需申报销售额合计", lambda report: report.total_sales),
            ),
            card_metrics=(
                rule_group_total("德国发出欧盟销售额", "德国发出欧盟"),
                rule_group_total("平台遗漏订单(德国境内,AE=0)", "德国境内遗漏"),
                rule_group_total("平台遗漏订单(目的国德国,AE非零)", "目的国德国遗漏"),
                derived_metric("Seller需申报合计", lambda report: report.total_sales),
            ),
        ),
        "it": CountryConfig(
            code="it",
            slug="italy",
            name_zh="意大利",
            title="意大利季度申报税金计算",
            description="按意大利算税逻辑独立汇总 B1、B2、F 三段销售额，并计算净销售额与销项税。",
            sales_report_label="意大利销售报告 CSV",
            logic_doc_label="意大利税金逻辑 PDF",
            logic_doc_accept=".pdf,application/pdf",
            emblem_path="/emblem_it.svg",
            excluded_transaction_types=("COMMINGLING BUY",),
            rules=(
                RuleSpec(
                    rule_id="IT_B1",
                    logic_group="B1销售额",
                    logic_bucket="CQ=SELLER_AE非零_CA非空_BP=IT_BQ欧盟",
                    description="CQ=SELLER；AE 非空且非 0；CA 非空；BP=IT；BQ 属于欧盟国家；BA 求和。",
                    matcher=it_rule_b1,
                ),
                RuleSpec(
                    rule_id="IT_B2",
                    logic_group="B2销售额",
                    logic_bucket="CQ=SELLER_AE非零_CA空白_BP欧盟_BQ=IT",
                    description="CQ=SELLER；AE 非空且非 0；CA 空白；BP 属于欧盟国家；BQ=IT；BA 求和。",
                    matcher=it_rule_b2,
                ),
                RuleSpec(
                    rule_id="IT_F",
                    logic_group="F销售额",
                    logic_bucket="CQ=SELLER_AE空白_CA空白_BP欧盟_BQ=IT",
                    description="CQ=SELLER；AE 为空白；CA 空白；BP 属于欧盟国家；BQ=IT；BA 求和。",
                    matcher=it_rule_f,
                ),
            ),
            summary_metrics=(
                rule_group_total("B1销售额"),
                rule_group_total("B2销售额"),
                rule_group_total("F销售额"),
                derived_metric("自行缴纳订单销售额合计(含税)", lambda report: report.total_sales),
                derived_metric("自行缴纳订单净销售额", lambda report: report.total_sales / IT_TAX_DIVISOR),
                derived_metric("销项税", lambda report: report.total_sales / IT_TAX_DIVISOR * IT_TAX_RATE),
            ),
            card_metrics=(
                rule_group_total("B1销售额", "B1"),
                rule_group_total("B2销售额", "B2"),
                rule_group_total("F销售额", "F"),
                derived_metric("自缴合计", lambda report: report.total_sales),
                derived_metric("净销售额", lambda report: report.total_sales / IT_TAX_DIVISOR),
                derived_metric("销项税", lambda report: report.total_sales / IT_TAX_DIVISOR * IT_TAX_RATE),
            ),
        ),
    }
)


def get_country_config(country_code: str) -> CountryConfig:
    try:
        return COUNTRY_CONFIGS[country_code]
    except KeyError as exc:
        raise ValueError(f"不支持的国家代码: {country_code}") from exc


def match_country_rule(
    row: dict[str, str],
    ctx: RuleContext,
    country: CountryConfig,
) -> MatchResult | None:
    for rule in country.rules:
        if rule.matcher(row, ctx):
            return MatchResult(
                rule_id=rule.rule_id,
                logic_group=rule.logic_group,
                logic_bucket=rule.logic_bucket,
                description=rule.description,
            )
    return None


def iter_matched_rows(csv_path: Path, country: CountryConfig) -> list[MatchedRow]:
    matched_rows: list[MatchedRow] = []
    ctx = RuleContext(country_code=country.code)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for source_row_number, row in enumerate(reader, start=2):
            match = match_country_rule(row, ctx, country)
            if not match:
                continue

            matched_rows.append(
                MatchedRow(
                    source_row_number=source_row_number,
                    rule_id=match.rule_id,
                    logic_group=match.logic_group,
                    logic_bucket=match.logic_bucket,
                    rule_description=match.description,
                    transaction_event_id=row.get("TRANSACTION_EVENT_ID", ""),
                    activity_period=row.get("ACTIVITY_PERIOD", ""),
                    marketplace=row.get("MARKETPLACE", ""),
                    transaction_type=row.get("TRANSACTION_TYPE", ""),
                    tax_collection_responsibility=row.get("TAX_COLLECTION_RESPONSIBILITY", ""),
                    tax_reporting_scheme=row.get("TAX_REPORTING_SCHEME", ""),
                    seller_sku=row.get("SELLER_SKU", ""),
                    asin=row.get("ASIN", ""),
                    item_description=row.get("ITEM_DESCRIPTION", ""),
                    qty=row.get("QTY", ""),
                    total_activity_value_amt_vat_incl=decimal_or_zero(
                        row.get("TOTAL_ACTIVITY_VALUE_AMT_VAT_INCL", "")
                    ),
                    transaction_currency_code=row.get("TRANSACTION_CURRENCY_CODE", ""),
                    price_of_items_vat_rate_percent=row.get("PRICE_OF_ITEMS_VAT_RATE_PERCENT", ""),
                    buyer_vat_number=row.get("BUYER_VAT_NUMBER", ""),
                    seller_arrival_country_vat_number=row.get(
                        "SELLER_ARRIVAL_COUNTRY_VAT_NUMBER", ""
                    ),
                    sale_depart_country=row.get("SALE_DEPART_COUNTRY", ""),
                    sale_arrival_country=row.get("SALE_ARRIVAL_COUNTRY", ""),
                    taxable_jurisdiction=row.get("TAXABLE_JURISDICTION", ""),
                    vat_calculation_imputation_country=row.get(
                        "VAT_CALCULATION_IMPUTATION_COUNTRY", ""
                    ),
                    invoice_url=row.get("INVOICE_URL", ""),
                )
            )
    return matched_rows


def summarize_by_group(matched_rows: list[MatchedRow]) -> dict[str, Decimal]:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for row in matched_rows:
        totals[row.logic_group] += row.total_activity_value_amt_vat_incl
    return dict(totals)


def summarize_by_rule(
    matched_rows: list[MatchedRow],
    country: CountryConfig,
) -> dict[str, dict[str, object]]:
    summary: dict[str, dict[str, object]] = {}
    for rule in country.rules:
        summary[rule.rule_id] = {
            "logic_group": rule.logic_group,
            "logic_bucket": rule.logic_bucket,
            "description": rule.description,
            "count": 0,
            "amount": Decimal("0"),
        }

    for row in matched_rows:
        bucket = summary[row.rule_id]
        bucket["count"] = int(bucket["count"]) + 1
        bucket["amount"] = Decimal(bucket["amount"]) + row.total_activity_value_amt_vat_incl
    return summary


def build_report_data(country: CountryConfig, matched_rows: list[MatchedRow]) -> CountryReportData:
    group_totals = summarize_by_group(matched_rows)
    total_sales = sum(
        (row.total_activity_value_amt_vat_incl for row in matched_rows),
        start=Decimal("0"),
    )
    sale_only_total = sum(
        (
            row.total_activity_value_amt_vat_incl
            for row in matched_rows
            if row.transaction_type.strip().upper() == "SALE"
        ),
        start=Decimal("0"),
    )
    return CountryReportData(
        country=country,
        matched_rows=matched_rows,
        group_totals=group_totals,
        total_sales=total_sales,
        sale_only_total=sale_only_total,
    )


def evaluate_metrics(report: CountryReportData, specs: tuple[MetricSpec, ...]) -> list[tuple[str, Decimal]]:
    return [(spec.label, spec.calculator(report)) for spec in specs]


def extract_currency_codes(matched_rows: list[MatchedRow]) -> list[str]:
    return sorted({row.transaction_currency_code for row in matched_rows if row.transaction_currency_code})


def autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def build_summary_sheet(
    wb: Workbook,
    report: CountryReportData,
    csv_path: Path,
    logic_doc_path: Path | None,
) -> None:
    ws = wb.active
    ws.title = "汇总"

    currency_codes = extract_currency_codes(report.matched_rows)

    ws.append(("指标", "值"))
    base_rows: list[tuple[str, object]] = [
        ("国家", report.country.name_zh),
        ("输入销售报告", str(csv_path)),
        ("税金逻辑来源", str(logic_doc_path) if logic_doc_path else "未提供"),
        ("命中规则记录数", len(report.matched_rows)),
    ]
    for row in base_rows:
        ws.append(row)

    for label, amount in evaluate_metrics(report, report.country.summary_metrics):
        ws.append((label, float(amount)))

    ws.append(("币种", ", ".join(currency_codes) if currency_codes else ""))
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    for row_index in range(2, ws.max_row + 1):
        if isinstance(ws.cell(row=row_index, column=2).value, (int, float)):
            ws.cell(row=row_index, column=2).number_format = "0.00"

    autosize_worksheet(ws)
    ws.freeze_panes = "A2"


def build_rule_sheet(
    wb: Workbook,
    matched_rows: list[MatchedRow],
    country: CountryConfig,
) -> None:
    ws = wb.create_sheet("规则说明")
    ws.append(
        [
            "rule_id",
            "logic_group",
            "logic_bucket",
            "matched_rows",
            "matched_amount",
            "description",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rule_summary = summarize_by_rule(matched_rows, country)
    for rule in country.rules:
        ws.append(
            [
                rule.rule_id,
                rule.logic_group,
                rule.logic_bucket,
                rule_summary[rule.rule_id]["count"],
                float(Decimal(rule_summary[rule.rule_id]["amount"])),
                rule.description,
            ]
        )

    for row_index in range(2, ws.max_row + 1):
        ws.cell(row=row_index, column=5).number_format = "0.00"

    autosize_worksheet(ws)
    ws.freeze_panes = "A2"


def build_detail_sheet(wb: Workbook, matched_rows: list[MatchedRow]) -> None:
    ws = wb.create_sheet("命中明细")
    headers = [
        "source_row_number",
        "rule_id",
        "logic_group",
        "logic_bucket",
        "rule_description",
        "transaction_event_id",
        "activity_period",
        "marketplace",
        "transaction_type",
        "tax_collection_responsibility",
        "tax_reporting_scheme",
        "seller_sku",
        "asin",
        "item_description",
        "qty",
        "total_activity_value_amt_vat_incl",
        "transaction_currency_code",
        "price_of_items_vat_rate_percent",
        "buyer_vat_number",
        "seller_arrival_country_vat_number",
        "sale_depart_country",
        "sale_arrival_country",
        "taxable_jurisdiction",
        "vat_calculation_imputation_country",
        "invoice_url",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in matched_rows:
        ws.append(
            [
                row.source_row_number,
                row.rule_id,
                row.logic_group,
                row.logic_bucket,
                row.rule_description,
                row.transaction_event_id,
                row.activity_period,
                row.marketplace,
                row.transaction_type,
                row.tax_collection_responsibility,
                row.tax_reporting_scheme,
                row.seller_sku,
                row.asin,
                row.item_description,
                row.qty,
                float(row.total_activity_value_amt_vat_incl),
                row.transaction_currency_code,
                row.price_of_items_vat_rate_percent,
                row.buyer_vat_number,
                row.seller_arrival_country_vat_number,
                row.sale_depart_country,
                row.sale_arrival_country,
                row.taxable_jurisdiction,
                row.vat_calculation_imputation_country,
                row.invoice_url,
            ]
        )

    amount_column = headers.index("total_activity_value_amt_vat_incl") + 1
    for row_index in range(2, ws.max_row + 1):
        ws.cell(row=row_index, column=amount_column).number_format = "0.00"

    autosize_worksheet(ws)
    ws.freeze_panes = "A2"


def build_workbook(
    report: CountryReportData,
    csv_path: Path,
    logic_doc_path: Path | None,
) -> Workbook:
    wb = Workbook()
    build_summary_sheet(wb, report, csv_path, logic_doc_path)
    build_rule_sheet(wb, report.matched_rows, report.country)
    build_detail_sheet(wb, report.matched_rows)
    return wb


def generate_self_tax_report(
    csv_path: Path,
    country_code: str,
    logic_pdf_path: Path | None = None,
    output_path: Path | None = None,
) -> dict[str, object]:
    country = get_country_config(country_code)
    csv_path = csv_path.expanduser().resolve()
    logic_doc_path = logic_pdf_path.expanduser().resolve() if logic_pdf_path else None

    if not csv_path.exists():
        raise FileNotFoundError(f"未找到销售报告: {csv_path}")

    if output_path is None:
        output_path = csv_path.with_name(f"{csv_path.stem}_{country.name_zh}税金汇总.xlsx")
    output_path = output_path.expanduser().resolve()

    matched_rows = iter_matched_rows(csv_path, country)
    report = build_report_data(country, matched_rows)
    workbook = build_workbook(report, csv_path, logic_doc_path)
    workbook.save(output_path)

    summary_metrics = evaluate_metrics(report, country.summary_metrics)
    card_metrics = evaluate_metrics(report, country.card_metrics)
    currency_codes = extract_currency_codes(report.matched_rows)

    return {
        "country_code": country.code,
        "country_name": country.name_zh,
        "country_slug": country.slug,
        "output_path": output_path,
        "matched_rows": matched_rows,
        "row_count": len(matched_rows),
        "group_totals": report.group_totals,
        "summary_metrics": summary_metrics,
        "card_metrics": card_metrics,
        "currency_codes": currency_codes,
        "total_sales": report.total_sales,
    }


def main() -> None:
    args = parse_args()
    result = generate_self_tax_report(
        csv_path=args.csv_path,
        country_code=args.country,
        logic_pdf_path=args.logic_pdf_path,
        output_path=args.output_path,
    )

    print(f"国家: {result['country_name']}")
    print(f"输出文件: {result['output_path']}")
    print(f"命中规则记录数: {result['row_count']}")
    for label, amount in result["summary_metrics"]:
        print(f"{label}: {amount:.2f}")


if __name__ == "__main__":
    main()
